from openpyxl import load_workbook
from openpyxl.workbook import Workbook
import inspect
import os


def excel_write(file_path, wb_object, sheet_object, user, pwd):
    user_list = [user, pwd]
    max_row = sheet_object.max_row
    max_column = sheet_object.max_column

    for col in range(1, max_column + 1):
        cell = sheet_object.cell(row=max_row + 1, column=col)
        cell.value = user_list[col - 1]

    wb_object.save(file_path)


def excel_load():
    file_path = os.path.join('files', 'db.xlsx')
    if os.path.exists(file_path):
        wb_object = load_workbook(file_path)
        sheet_object = wb_object.worksheets[0]
    else:
        wb_object = Workbook()
        sheet_object = wb_object.worksheets[0]
        param_count = len(inspect.signature(excel_write).parameters) - 3
        title_list = ['用户名', '密码']
        for i in range(param_count):
            sheet_object.cell(row=1, column=i + 1).value = title_list[i]

    return file_path, wb_object, sheet_object


def run():
    file_path, wb_object, sheet_object = excel_load()
    while True:
        user = input(">>>用户名(Q/q)：")
        if user.upper() == "Q":
            break
        pwd = input(">>>密码：")
        excel_write(file_path, wb_object, sheet_object, user, pwd)


if __name__ == '__main__':
    run()