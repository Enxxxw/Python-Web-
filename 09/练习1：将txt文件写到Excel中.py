from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
import os


def file_get(file_path):
    txt_dict = {}
    with open(file_path, 'r', encoding='utf-8') as f:
        for num, line in enumerate(f.readlines(), 1):
            line = line.strip()
            line_list = line.split(',')
            txt_dict[num] = line_list
    # print(txt_dict)
    return txt_dict


def cell_type(cell, border_color, font_color, font_size, font_name,fill_color):
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        top=Side(style='medium', color=border_color),
        right=Side(style='medium', color=border_color),
        left=Side(style='medium', color=border_color),
        bottom=Side(style='medium', color=border_color),
    )
    cell.font = Font(color=font_color, size=font_size, name=font_name)
    cell.fill = PatternFill("solid", fgColor=fill_color)


def xlsx_write(txt_dict, title_list):
    wb = Workbook()
    sheet = wb.worksheets[0]
    sheet.title = "user"

    for row, text in enumerate(txt_dict.values(), 1):
        if row == 1:
            height = 80
        else:
            height = 40
        sheet.row_dimensions[row].height = height
        for i in range(1, len(text)+1):
            sheet.column_dimensions[chr(ord('A')+i-1)].width = 50
            cell = sheet.cell(row=row, column=i)
            if row == 1:
                cell.value = "中国联通"
                cell_type(cell, '000000', 'FFFFFF', 50, '仿宋', '4169E1')
                sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
                break
            elif row == 2:
                cell.value = title_list[i - 1]
                cell_type(cell, '000000', 'FFFFFF', 30, '仿宋', '4169E1')
            else:
                cell.value = text[i - 1]
                cell_type(cell, '000000', '000000', 30, '仿宋', '87CEFA')

    wb.save('files/p12.xlsx')

def run():
    title_list = ['用户', '密码', '邮箱']
    txt_path = os.path.join('files', 'user.txt')
    txt_dict = file_get(txt_path)
    xlsx_write(txt_dict, title_list)

if __name__ == '__main__':
    run()