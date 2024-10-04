from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment, GradientFill, Font, PatternFill

wb = Workbook()
sheet0 = wb.worksheets[0]

cell = sheet0.cell(row=2, column=2)
cell.value = "我是中国人"
cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
# 填充背景色
cell.font = Font(color='FF0000', size=40, name='仿宋')
cell.fill = PatternFill("solid", fgColor="F4A460")

wb.save('files/p9.xlsx')