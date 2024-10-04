from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

wb = Workbook()
sheet0 = wb.worksheets[0]

sheet0.row_dimensions[1].height = 50
sheet0.column_dimensions['A'].width = 30
sheet0.column_dimensions['B'].width = 30
sheet0.column_dimensions['C'].width = 30
sheet0.column_dimensions['D'].width = 30
sheet0.column_dimensions['E'].width = 30

# 居中和自动换行
name_list = ["张三", "李四", "马五", "牛六", "赵七"]
for col, text in enumerate(name_list, 1):
    cell = sheet0.cell(1, col)
    cell.value = text
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        top=Side(style='medium', color='000000'),
        right=Side(style='medium', color='0000FF'),
        left=Side(style='medium', color='0000FF'),
        bottom=Side(style='medium', color='000000'),
    )
    cell.font = Font(color='ff0000', size=30, name='仿宋')
    cell.fill = PatternFill("solid", fgColor="F4A460")

wb.save('files/p12.xlsx')